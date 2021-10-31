import argparse
import os
from sys import exit

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from PIL import Image


def rgb_to_hex(r, g, b):
    """
    Convert RGB color to an Hexadecimal representation
    """
    return "%02x%02x%02x" % (r, g, b)


def verify_xlsx_ext(file_name):
    """
    Verify if the file_name has the right extension
    """
    if not file_name.endswith(".xlsx"):
        if '.' in file_name:
            file_name = file_name.split('.')[0]

        file_name += ".xlsx"

    return file_name


def paint_it(img_name, file_name):

    """
    Just paint the spreadsheet with the colors
    """

    try:
        img = Image.open(img_name)

    except FileNotFoundError:
        print("Invalid path for your image")
        exit(2)

    # Force the image to scaled into 128x128
    # This prevent the death of your machine
    img.thumbnail((128, 128), Image.ANTIALIAS)

    img = np.array(img)

    # For numpy strings are objects, so we create an empty
    # numpy array of type object
    new_array = np.empty((img.shape[0], img.shape[1]), object)

    # Place the hexadecimal number on each cell
    for ind_row, rows in enumerate(img):
        for ind_col, cols in enumerate(rows):
            rgb = img[ind_row, ind_col, :]
            new_array[ind_row, ind_col] = rgb_to_hex(rgb[0], rgb[1], rgb[2])

    # Yes, I imported Pandas to convert a Numpy array to a Pandas DataFrame
    # for save it as a excel ....
    df = pd.DataFrame(new_array)
    df.to_excel("tmp.xlsx", header=False, index=False)

    wb = openpyxl.load_workbook("tmp.xlsx")
    ws = wb.active

    # Zoom out
    ws.sheet_view.zoomScale = 10

    # Now I read the tmp.xlsx and convert the hexadecimal numbers into the color of the cell
    # and remove the value from it
    for rows in ws.iter_rows(min_row=1, max_row=new_array.shape[0], min_col=1):
        for cell in rows:
            hex_color = "{}".format(str(cell.value).upper())
            cell.fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
            cell.value = ""

    # Make column with small width
    for cols in range(1, new_array.shape[1] + 1):
        j = openpyxl.utils.get_column_letter(cols)
        ws.column_dimensions[j].width = 1

    # Make row with small height
    for i in range(1, new_array.shape[0] + 1):
        ws.row_dimensions[i].height = 10

    os.remove("tmp.xlsx")

    # Ensure that the extension is right
    file_name = verify_xlsx_ext(file_name)

    wb.save(file_name)


def main():
	"""
		Main function where it gets command line's arguments
	"""
    parser = argparse.ArgumentParser(
        description="Let's draw a photo on a spreadsheet"
    )

    parser.add_argument("-i", help="Image name", type=str, dest="img_name")
    parser.add_argument("-f", help="File name", type=str, dest="file_name")

    args = parser.parse_args()

    if (not args.img_name) or (not args.file_name):
        parser.print_help()
        exit(1)

    paint_it(args.img_name, args.file_name)


if __name__ == "__main__":
    main()
